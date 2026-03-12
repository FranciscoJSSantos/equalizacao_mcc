import streamlit as st
import pandas as pd
import openpyxl
import io
import os
import sys

st.set_page_config(page_title="Equalização MCC", layout="wide")

st.write("✅ App carregado")

try:
    from auth import require_auth, logout
    st.write("✅ auth.py importado")
    user = require_auth()
except Exception as _e:
    st.error(f"Erro: {_e}")
    import traceback
    st.code(traceback.format_exc())
    st.stop()

with st.sidebar:
    st.markdown(f"**{user['name']}**")
    st.caption(user["email"])
    st.markdown("---")
    if st.button("Sair", use_container_width=True):
        logout()

st.title("Equalização de Fornecedores - MCC")

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

class ParseError(Exception):
    pass


def parse_excel(file) -> pd.DataFrame:
    """Parse ROTINA and PG sheets into a long-format DataFrame."""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        raise ParseError("Não foi possível abrir o arquivo. Certifique-se de que é um Excel (.xlsx) válido.")

    sheets_found = [s for s in ["ROTINA", "PG"] if s in wb.sheetnames]
    if not sheets_found:
        raise ParseError(
            f"Nenhuma sheet 'ROTINA' ou 'PG' encontrada. "
            f"Sheets disponíveis no arquivo: {', '.join(wb.sheetnames)}"
        )

    records = []
    for tipo in sheets_found:
        ws = wb[tipo]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ParseError(f"A sheet '{tipo}' está vazia.")

        header = rows[0]
        if len(header) < 4:
            raise ParseError(
                f"A sheet '{tipo}' deve ter pelo menos 4 colunas "
                f"(código, função, unidade, e ao menos um fornecedor). "
                f"Encontradas: {len(header)} coluna(s)."
            )

        expected = ["código", "função", "unidade"]
        for i, exp in enumerate(expected):
            val = str(header[i]).strip().lower() if header[i] else ""
            if exp not in val:
                raise ParseError(
                    f"Sheet '{tipo}': coluna {i+1} deveria ser '{exp}', "
                    f"mas encontrou '{header[i]}'. Verifique o cabeçalho."
                )

        supplier_cols = []
        for h in header[3:]:
            if h is None:
                supplier_cols.append(None)
                continue
            parts = str(h).strip().split()
            if len(parts) < 2:
                raise ParseError(
                    f"Sheet '{tipo}': coluna de fornecedor '{h}' deve ter o formato "
                    f"'FORNECEDOR CENTRO' (ex: 'LATIN JAC')."
                )
            supplier_cols.append((parts[0], parts[-1]))

        data_rows = [r for r in rows[1:] if r[0] is not None]
        if not data_rows:
            raise ParseError(f"A sheet '{tipo}' não possui linhas de dados (apenas cabeçalho).")

        for row in data_rows:
            codigo = row[0]
            funcao = row[1]
            unidade = row[2]
            for i, sc in enumerate(supplier_cols):
                if sc is None:
                    continue
                val = row[3 + i] if 3 + i < len(row) else None
                if val is None:
                    continue
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    continue
                records.append({
                    "tipo": tipo,
                    "codigo": codigo,
                    "funcao": funcao,
                    "unidade": unidade,
                    "fornecedor": sc[0],
                    "centro": sc[1],
                    "valor": val,
                })

    if not records:
        raise ParseError("O arquivo foi lido, mas nenhum valor numérico foi encontrado nas colunas de fornecedores.")

    return pd.DataFrame(records)


def highlight_min(s):
    is_min = s == s.min()
    return ["background-color: #c6efce; color: #276221; font-weight:bold" if v else "" for v in is_min]


def highlight_min_col(df_styled, col):
    """Highlight the row with minimum value in a given column."""
    min_val = df_styled.data[col].min()
    return df_styled.apply(
        lambda row: ["background-color: #c6efce; color: #276221; font-weight:bold"
                     if row[col] == min_val else "" for _ in row],
        axis=1
    )


# ─────────────────────────────────────────────────────────────
# 1. Upload
# ─────────────────────────────────────────────────────────────

st.header("1. Importar Planilha")
uploaded = st.file_uploader("Faça upload do arquivo Excel (.xlsx)", type=["xlsx"])

# Fallback to local file for demo
_BASE_DIR = sys._MEIPASS if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.join(_BASE_DIR, "sheet", "Cópia de MCC (version 3).xlsx")

df_raw = None
if uploaded:
    try:
        df_raw = parse_excel(uploaded)
        st.success(f"Planilha carregada: {len(df_raw)} registros em {df_raw['tipo'].nunique()} tipo(s) de serviço.")
    except ParseError as e:
        st.error(f"Erro ao ler a planilha: {e}")
        st.stop()
    except Exception as e:
        st.error(f"Erro inesperado: {e}")
        st.stop()
elif os.path.exists(DEFAULT_FILE):
    try:
        with open(DEFAULT_FILE, "rb") as f:
            df_raw = parse_excel(f)
        st.info("Usando arquivo padrão local: Cópia de MCC (version 3).xlsx")
    except ParseError as e:
        st.error(f"Erro no arquivo padrão: {e}")
        st.stop()

if df_raw is None or df_raw.empty:
    st.warning("Faça o upload de uma planilha para continuar.")
    st.stop()

tipos = sorted(df_raw["tipo"].unique())
centros = sorted(df_raw["centro"].unique())
funcoes = sorted(df_raw["funcao"].dropna().unique())
fornecedores = sorted(df_raw["fornecedor"].unique())

# ─────────────────────────────────────────────────────────────
# Tabs
# ─────────────────────────────────────────────────────────────

tab2, tab3, tab4, tab5 = st.tabs([
    "2. Comparativo",
    "3. Cherrypicking",
    "4. Should Cost",
    "5. Simulação",
])

# ─────────────────────────────────────────────────────────────
# 2. Comparativo
# ─────────────────────────────────────────────────────────────

with tab2:
    st.header("Comparativo Linha a Linha")
    c1, c2 = st.columns(2)
    tipo_sel = c1.selectbox("Tipo de Serviço", tipos, key="comp_tipo")
    centro_sel = c2.multiselect("Centro(s)", centros, default=centros, key="comp_centro")

    df_c = df_raw[(df_raw["tipo"] == tipo_sel) & (df_raw["centro"].isin(centro_sel))]

    if df_c.empty:
        st.warning("Sem dados para os filtros selecionados.")
    else:
        pivot = (
            df_c
            .pivot_table(
                index=["codigo", "funcao", "unidade"],
                columns=["fornecedor", "centro"],
                values="valor",
                aggfunc="first",
            )
        )
        pivot.columns = [f"{f} / {c}" for f, c in pivot.columns]
        pivot = pivot.reset_index()
        pivot.rename(columns={"codigo": "Cód.", "funcao": "Função", "unidade": "Unidade"}, inplace=True)

        value_cols = [c for c in pivot.columns if c not in ["Cód.", "Função", "Unidade"]]

        styled = pivot.style.apply(highlight_min, subset=value_cols, axis=1).format(
            {c: lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "-" for c in value_cols}
        )
        st.dataframe(styled, width="stretch", hide_index=True)

# ─────────────────────────────────────────────────────────────
# 3. Cherrypicking
# ─────────────────────────────────────────────────────────────

with tab3:
    st.header("Cherrypicking por Fornecedor e Centro")
    c1, c2 = st.columns(2)
    tipo_ck = c1.selectbox("Tipo de Serviço", tipos, key="ck_tipo")
    centro_ck = c2.multiselect("Centro(s)", centros, default=centros, key="ck_centro")

    df_ck = df_raw[(df_raw["tipo"] == tipo_ck) & (df_raw["centro"].isin(centro_ck))].copy()

    if df_ck.empty:
        st.warning("Sem dados para os filtros selecionados.")
    else:
        # For each (funcao, centro) find min valor and best supplier
        idx_min = df_ck.groupby(["funcao", "centro"])["valor"].idxmin()
        df_best = df_ck.loc[idx_min].copy()
        df_best = df_best.rename(columns={
            "funcao": "Função", "centro": "Centro",
            "fornecedor": "Melhor Fornecedor", "valor": "Melhor Preço",
            "unidade": "Unidade", "codigo": "Cód."
        })

        st.subheader("Melhores preços (cherrypicking)")
        st.dataframe(
            df_best[["Cód.", "Função", "Unidade", "Centro", "Melhor Fornecedor", "Melhor Preço"]]
            .sort_values(["Centro", "Cód."])
            .style.format({"Melhor Preço": "R$ {:,.2f}"}),
            width="stretch",
            hide_index=True,
        )

        st.subheader("Tabela completa com destaque do menor")
        pivot_ck = (
            df_ck
            .pivot_table(
                index=["codigo", "funcao", "unidade", "centro"],
                columns="fornecedor",
                values="valor",
                aggfunc="first",
            )
            .reset_index()
        )
        pivot_ck.rename(columns={"codigo": "Cód.", "funcao": "Função", "unidade": "Unidade", "centro": "Centro"}, inplace=True)
        forn_cols = [c for c in pivot_ck.columns if c not in ["Cód.", "Função", "Unidade", "Centro"]]

        styled_ck = pivot_ck.style.apply(highlight_min, subset=forn_cols, axis=1).format(
            {c: lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "-" for c in forn_cols}
        )
        st.dataframe(styled_ck, width="stretch", hide_index=True)

# ─────────────────────────────────────────────────────────────
# 4. Should Cost
# ─────────────────────────────────────────────────────────────

with tab4:
    st.header("Should Cost — Melhor Preço Total por Fornecedor e Centro")
    tipo_sc = st.selectbox("Tipo de Serviço", tipos, key="sc_tipo")

    df_sc = df_raw[df_raw["tipo"] == tipo_sc].copy()

    # For each (fornecedor, centro): count items covered + sum of prices
    agg = (
        df_sc
        .groupby(["fornecedor", "centro"])
        .agg(
            total_itens=("valor", "count"),
            preco_total=("valor", "sum"),
            preco_medio=("valor", "mean"),
        )
        .reset_index()
    )

    # Total items per centro (across all suppliers via cherrypicking)
    total_por_centro = df_sc.groupby("centro")["funcao"].nunique().reset_index()
    total_por_centro.columns = ["centro", "total_funcoes"]
    agg = agg.merge(total_por_centro, on="centro", how="left")
    agg["cobertura_pct"] = (agg["total_itens"] / agg["total_funcoes"] * 100).round(1)

    # Best supplier per centro
    best_per_centro = agg.loc[agg.groupby("centro")["preco_total"].idxmin(), ["centro", "fornecedor"]].copy()
    best_per_centro.columns = ["centro", "melhor_fornecedor"]
    agg = agg.merge(best_per_centro, on="centro", how="left")
    agg["e_melhor"] = agg["fornecedor"] == agg["melhor_fornecedor"]

    for centro_name in sorted(agg["centro"].unique()):
        df_c = agg[agg["centro"] == centro_name].sort_values("preco_total").copy()
        df_c = df_c.drop(columns=["melhor_fornecedor"])

        st.subheader(f"Centro: {centro_name}")

        def row_color(row):
            if row["e_melhor"]:
                return ["background-color: #c6efce; color: #276221; font-weight:bold"] * len(row)
            return [""] * len(row)

        display_cols = ["fornecedor", "centro", "total_itens", "cobertura_pct", "preco_total", "preco_medio"]
        rename_map = {
            "fornecedor": "Fornecedor", "centro": "Centro",
            "total_itens": "Itens", "cobertura_pct": "Cobertura (%)",
            "preco_total": "Preço Total (R$)", "preco_medio": "Preço Médio (R$)",
        }
        df_show = df_c[display_cols + ["e_melhor"]].rename(columns=rename_map)

        styled_sc = (
            df_show.style
            .apply(row_color, axis=1)
            .format({
                "Preço Total (R$)": "R$ {:,.2f}",
                "Preço Médio (R$)": "R$ {:,.2f}",
                "Cobertura (%)": "{:.1f}%",
            })
            .hide(axis="columns", subset=["e_melhor"])
        )
        st.dataframe(styled_sc, width="stretch", hide_index=True)

        all_funcoes_centro = df_sc[df_sc["centro"] == centro_name][["funcao", "unidade"]].drop_duplicates("funcao")
        n_total_centro = len(all_funcoes_centro)
        for _, frow in df_c.sort_values("preco_total").iterrows():
            forn = frow["fornecedor"]
            forn_data = df_sc[(df_sc["centro"] == centro_name) & (df_sc["fornecedor"] == forn)][["funcao", "valor"]]
            n_cov = len(forn_data)
            label = f"{forn}  —  {n_cov}/{n_total_centro} itens  |  R$ {frow['preco_total']:,.2f}"
            with st.expander(label):
                detail_rows = []
                for _, fn_row in all_funcoes_centro.sort_values("funcao").iterrows():
                    fn = fn_row["funcao"]
                    match = forn_data[forn_data["funcao"] == fn]
                    if not match.empty:
                        detail_rows.append({"Função": fn, "Unidade": fn_row["unidade"], "Preço (R$)": match.iloc[0]["valor"], "Coberto": "✓"})
                    else:
                        detail_rows.append({"Função": fn, "Unidade": fn_row["unidade"], "Preço (R$)": None, "Coberto": "✗"})
                df_det = pd.DataFrame(detail_rows)
                def _det_color(row):
                    return ["color: #999999"] * len(row) if row["Coberto"] == "✗" else [""] * len(row)
                st.dataframe(
                    df_det.style.apply(_det_color, axis=1).format({"Preço (R$)": lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "N/D"}),
                    hide_index=True,
                )

# ─────────────────────────────────────────────────────────────
# 5. Simulação
# ─────────────────────────────────────────────────────────────

with tab5:
    st.header("Simulação")
    c1, c2 = st.columns(2)
    tipo_sim = c1.selectbox("Tipo de Serviço", tipos, key="sim_tipo")
    centro_sim = c2.selectbox("Centro", centros, key="sim_centro")

    df_sim_base = df_raw[(df_raw["tipo"] == tipo_sim) & (df_raw["centro"] == centro_sim)].copy()
    funcoes_centro = df_sim_base["funcao"].dropna().unique().tolist()

    st.subheader("Selecione as funções e quantidades")
    funcoes_sel = st.multiselect("Funções", sorted(funcoes_centro), key="sim_funcoes")

    if not funcoes_sel:
        st.info("Selecione pelo menos uma função para continuar.")
    else:
        qtds = {}
        cols_q = st.columns(min(len(funcoes_sel), 4))
        for i, fn in enumerate(funcoes_sel):
            un_row = df_sim_base[df_sim_base["funcao"] == fn]["unidade"].iloc[0] if not df_sim_base[df_sim_base["funcao"] == fn].empty else ""
            label = f"{fn[:30]} ({un_row})"
            qtds[fn] = cols_q[i % len(cols_q)].number_input(label, min_value=0.0, value=1.0, step=1.0, key=f"q_{fn}")

        df_sim = df_sim_base[df_sim_base["funcao"].isin(funcoes_sel)].copy()
        df_sim["quantidade"] = df_sim["funcao"].map(qtds)
        df_sim["total"] = df_sim["valor"] * df_sim["quantidade"]

        # --- Cherrypicking simulado ---
        st.subheader("Cherrypicking (melhor fornecedor por função)")
        idx_best = df_sim.groupby("funcao")["valor"].idxmin()
        df_cherry = df_sim.loc[idx_best, ["codigo", "funcao", "unidade", "fornecedor", "valor", "quantidade", "total"]].copy()
        df_cherry.rename(columns={
            "codigo": "Cód.", "funcao": "Função", "unidade": "Unidade",
            "fornecedor": "Melhor Fornecedor", "valor": "Preço Unit.",
            "quantidade": "Qtd", "total": "Total",
        }, inplace=True)

        total_cherry = df_cherry["Total"].sum()
        st.dataframe(
            df_cherry.style.format({"Preço Unit.": "R$ {:,.2f}", "Total": "R$ {:,.2f}", "Qtd": "{:,.1f}"}),
            use_container_width=True,
            hide_index=True,
        )
        st.metric("Total Cherrypicking", f"R$ {total_cherry:,.2f}")

        # --- Should Cost simulado ---
        st.subheader("Should Cost por Fornecedor (pacote completo)")
        agg_sim = (
            df_sim
            .groupby("fornecedor")
            .agg(
                funcoes_cobertas=("funcao", "nunique"),
                total_custo=("total", "sum"),
            )
            .reset_index()
        )
        agg_sim["cobre_tudo"] = agg_sim["funcoes_cobertas"] == len(funcoes_sel)
        agg_sim = agg_sim.sort_values("total_custo")

        min_custo_completo = agg_sim[agg_sim["cobre_tudo"]]["total_custo"].min() if agg_sim["cobre_tudo"].any() else None

        def row_color_sim(row):
            cobre = row["Cobre Tudo?"]
            custo = row["Custo Total (R$)"]
            if cobre and min_custo_completo is not None and custo == min_custo_completo:
                return ["background-color: #c6efce; color: #276221; font-weight:bold"] * len(row)
            if not cobre:
                return ["color: #999999"] * len(row)
            return [""] * len(row)

        agg_sim_display = agg_sim.rename(columns={
            "fornecedor": "Fornecedor",
            "funcoes_cobertas": "Funções Cobertas",
            "total_custo": "Custo Total (R$)",
            "cobre_tudo": "Cobre Tudo?",
        })
        styled_sim = (
            agg_sim_display.style
            .apply(row_color_sim, axis=1)
            .format({"Custo Total (R$)": "R$ {:,.2f}"})
        )
        st.dataframe(styled_sim, width="stretch", hide_index=True)

        for _, frow in agg_sim.sort_values("total_custo").iterrows():
            forn = frow["fornecedor"]
            forn_sim_data = df_sim[df_sim["fornecedor"] == forn][["funcao", "unidade", "valor", "quantidade", "total"]]
            n_cov = len(forn_sim_data)
            cobre_label = "Cobre tudo" if frow["cobre_tudo"] else f"{n_cov}/{len(funcoes_sel)} itens"
            label = f"{forn}  —  {cobre_label}  |  R$ {frow['total_custo']:,.2f}"
            with st.expander(label):
                det_rows = []
                for fn in sorted(funcoes_sel):
                    match = forn_sim_data[forn_sim_data["funcao"] == fn]
                    if not match.empty:
                        r = match.iloc[0]
                        det_rows.append({"Função": fn, "Unidade": r["unidade"], "Qtd": r["quantidade"], "Preço Unit. (R$)": r["valor"], "Total (R$)": r["total"], "Coberto": "✓"})
                    else:
                        det_rows.append({"Função": fn, "Unidade": "", "Qtd": qtds.get(fn, 0), "Preço Unit. (R$)": None, "Total (R$)": None, "Coberto": "✗"})
                df_det_sim = pd.DataFrame(det_rows)
                def _det_sim_color(row):
                    return ["color: #999999"] * len(row) if row["Coberto"] == "✗" else [""] * len(row)
                st.dataframe(
                    df_det_sim.style.apply(_det_sim_color, axis=1).format({
                        "Preço Unit. (R$)": lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "N/D",
                        "Total (R$)": lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "N/D",
                        "Qtd": "{:,.1f}",
                    }),
                    hide_index=True,
                )

        best_full = agg_sim[agg_sim["cobre_tudo"]]
        if not best_full.empty:
            best_row = best_full.iloc[0]
            st.metric(
                f"Melhor fornecedor completo: {best_row['fornecedor']}",
                f"R$ {best_row['total_custo']:,.2f}",
                delta=f"vs cherrypicking: R$ {best_row['total_custo'] - total_cherry:,.2f}",
                delta_color="inverse",
            )
        else:
            st.warning("Nenhum fornecedor cobre todas as funções selecionadas neste centro.")
